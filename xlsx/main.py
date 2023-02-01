from openpyxl import load_workbook

wb = load_workbook("precos.xlsx")

# active worksheet -> sheet_wb = wb.active
sheet_wb = wb["precos"]

# capturing the contents of a cell
a1 = sheet_wb["A1"]
print(a1.value)

# or
a1 = sheet_wb.cell(row=1, column=1)
print(a1.value)
