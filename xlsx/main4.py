from openpyxl import load_workbook

# add worksheet
file_xlsx = "precos.xlsx"
wb = load_workbook(file_xlsx)
wb.create_sheet("cars")
wb.save(file_xlsx)

# remove
wb.remove(wb["cars"])
wb.save(file_xlsx)

# copy
origin = wb["cars"]
destiny = wb.copy_worksheet(origin)
destiny.title = "new cars"

wb.save(file_xlsx)
