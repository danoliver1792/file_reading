from openpyxl import load_workbook
from openpyxl.styles import Alignment

# changing cell alignment
wb = load_workbook("precos.xlsx")
sheet_wb = wb["precos"]

sheet_wb.unmerge_cells("A2:A7")
sheet_wb["B2"] = "Descricao do Produto"
sheet_wb["B2"].aligment = Alignment(horizontal="center")
wb.save("precos.xlsx")
