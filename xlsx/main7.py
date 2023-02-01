from openpyxl import load_workbook

# merge cells
wb = load_workbook("precos.xlsx")
sheet_wb = wb["cars"]
sheet_wb.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)

# or
sheet_wb.merge_cells("D2:D7")
wb.save("precos.xlsx")
